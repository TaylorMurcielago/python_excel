from django.http import HttpResponse
import openpyxl
import xlrd
import json
import xlwt
import csv
from openpyxl.styles import PatternFill

from openpyxl.styles import colors
import numpy as np
from openpyxl.styles import Font, Color
from django.views.decorators.csrf import csrf_exempt

# import  pandas  as pd
import  pandas  as pd
import xlsxwriter as xw
import datacompy 
from openpyxl import load_workbook
# import xlwings as xw
culumn1=''
culumn2='' 
@csrf_exempt
def contrast(request):
    json_data=json.loads(request.body)
    
    a1=json_data[0].get("a1")
    a2=json_data[0].get("a2")
    keyID=json_data[0].get("column1")
#     workbooka = xw.Book(r'/Users/tayli/Desktop/project/python/薪资1.xlsx')
#     workbookb = xw.Book(r'/Users/tayli/Desktop/project/python/薪资2.xlsx')
#     readbook = xlrd.open_workbook(r'/Users/tayli/Desktop/project/python/薪资1.xlsx')
#     workbook = xlrd.open_workbook('/Users/tayli/Desktop/project/python/薪资2.xlsx')
#     workbook1= xw.Workbook('/Users/tayli/Desktop/project/python/薪资12.xlsx')
#     workbook2= xw.Workbook('/Users/tayli/Desktop/project/python/薪资22.xlsx')
#     df1=pd.read_excel('/Users/tayli/Desktop/project/python/薪资1.xlsx')
#     df2=pd.read_excel('/Users/tayli/Desktop/project/python/薪资2.xlsx')
# 创建表
    new_workbook=xlwt.Workbook()
    # # 创建sheet
    worksheet=new_workbook.add_sheet('对比结果')
    if(json_data[0].get('oror')=='CSV'):
           wb_a = pd.read_csv(json_data[0].get("a1"),error_bad_lines=False)
           wb_b = pd.read_csv(json_data[0].get("a2"),error_bad_lines=False)
    else:
           wb_a = pd.read_excel(json_data[0].get("a1"))
           wb_b = pd.read_excel(json_data[0].get("a2")) 


    compare = datacompy.Compare( wb_a,wb_b,join_columns=keyID) 
#     file_name='/Users/tayli/Desktop/project/python/test1.xlsx' 
#     print(compare.report(sample_count=5000)) 
#     df1_unq_rows=compare.df1_unq_rows
#     df2_unq_rows=compare.df1_unq_rows 

#     book = load_workbook('/Users/tayli/Desktop/project/python/test1.xlsx')
#     writer = pd.ExcelWriter('/Users/tayli/Desktop/project/python/test1.xlsx', engine='openpyxl')
#     writer.book = book
#     compare.compare.report().to_excel(writer, 'sheet1')
#     compare.df1_unq_rows.to_excel(writer, 'sheet1') 
#     compare.df2_unq_rows.to_excel(writer, 'sheet2') 
#     writer.save()
#     writer.close()


#     wb_a.fillna(100) 
#     wb_b.fillna(100)
#     wb_a = openpyxl.load_workbook("/Users/tayli/Desktop/project/python/薪资11.csv") 
#     wb_b = openpyxl.load_workbook("/Users/tayli/Desktop/project/python/薪资22.csv")
#     notEquial=[]
#     for i,cellobj in enumerate(wb_a.worksheets[0]['B'],1): 
#     # if(wb_a.worksheets[0].value):
#     # print(cellobj.value)
#     # print(wb_a.worksheets[0]['B'][0].value)
#     # 查询b中的某一个伙伴号在a表是不是有,首先伙伴号相等
#     # if(cellobj.value==wb_b.worksheets[0]['B'][i].value):
#         for j,cellobjB in enumerate(wb_b.worksheets[0]['B'],1):
#             # print(cellobj.value,cellobjB)
#             if(cellobj.value==cellobjB.value):
#                 mm=(cellobjB.value,j)
#                 # 再去判断目标行的值是不是相等 
#                 if(wb_b.worksheets[0]['D'][j-1].value==wb_a.worksheets[0]['D'][i-1].value):
#                         print(wb_a.worksheets[0]['D'][i-1].value)
#                 else: 
#                          print('不相等'+wb_a.worksheets[0]['D'][i-1].value) 
#                          notEquial.append(wb_a.worksheets[0]['D'][i-1].value)
              #   if(wb_b.worksheets[0][j-1][0].value==wb_a.worksheets[0][i-1][0].value):
              #          print('ok')
              #   else:
              #          print('不相等'+wb_b.worksheets[0][j-1][0]) 
#     for():
       # function2指定两个excel的员工编号必须是指定列
#     for mmm,xx in enumerate(np.array(wb_a),1):
#            print(xx)
#            print(mmm)
#     if((np.array(wb_a)==np.array(wb_b)).all()):
#             print('ooooooo')
#     for nn1,yy1 in enumerate(wb_a['伙伴工号'],0):
#            for nn2,yy2 in enumerate(wb_b['伙伴工号'],0):
                  
                 
#                   if(yy1==yy2):
#                          for a,coo in enumerate(wb_a.columns,0):
#                             # print(wb_a.iloc[nn1].fillna("我是填充")[coo])
#                             # print(wb_b.iloc[nn2].fillna("我是填充")[coo])
#                             if(wb_a.iloc[nn1].fillna("我是填充")[coo]!=wb_b.iloc[nn2].fillna("我是填充")[coo]):
#                                    print(wb_a.iloc[nn1].fillna("我是填充")[coo]) 
#                                    print(wb_b.iloc[nn2].fillna("我是填充")[coo])    

#     for ii,cellobj1 in enumerate(wb_a['伙伴工号'],0): 
#            for jj,cellobjB1 in enumerate(wb_b['伙伴工号'],0):
#                   if(cellobj1==cellobjB1):
#                          for a,coo in enumerate(wb_a.columns,0):
#                                  wb_a.iloc[ii].fillna("我是填充")
#                                  wb_b.iloc[jj].fillna("我是填充")
                                 
#                                  if(wb_a.iloc[ii].fillna("我是填充")[coo]!=wb_b.iloc[jj].fillna("我是填充")[coo]):
#                                         print(wb_a.iloc[ii][coo])
#                                         print(ii)
#                                         print(coo)
#                                         print(wb_b.iloc[jj][coo])
#                                         print(jj)
#                                         print(coo)

#                                         wb_a.iloc[ii].fillna("我是填充")[coo].font = Font(color=colors.BLACK, italic=True ,bold = True)
#                                         wb_a.iloc[ii].fillna("我是填充")[coo].fill = PatternFill("solid", fgColor="d71345") 
#                             # #  把对比出的结果写入一个附件excel
#                                         worksheet.write(jj,0,wb_a.iloc[ii].fillna("我是填充")[coo])
#                             # #  保存加过对比色文件
#                                         wb_a.save('/Users/tayli/Desktop/project/python/重标记色全文本对比结果备份.xlsx') 
#                             # # 保存晒出掉一样数据的结果，存入excel
#                                         new_workbook.save('/Users/tayli/Desktop/project/python/只保留不同值excel.xlsx')
                     #     if(wb_a.iloc[ii]['伙伴姓名']!=wb_b.iloc[jj]['伙伴姓名']):
                     #     获取到伙伴号相等的时候，两个a，b表中的行号。那么对比当前行所有的列是否相等即可
                     #  wb_a.worksheet[0][行][列]
                     # for mm,cellobj2 in enumerate(wb_a.worksheets[0][ii],1):
                     #        print(cellobj2)
                     #        print(mm)
                     # for mm,cellobjB2 in enumerate(wb_b.worksheets[0][jj],1):
                     #        print(cellobjB2)
                     
                     # print(cellobj1)
                     # for x,a in enumerate(wb_a.iloc[ii:1,],1):
                     # #     print(wb_a.worksheets[0][ii][a])
                     # #     print(wb_b.worksheets[0][jj][a])
                     #     if(a!=wb_b.iloc[jj:1,]):
                     # #            print('')
                     # #     else:
                     #            print(wb_a.worksheets[0][ii][a].value)
                     #            print(wb_b.worksheets[0][jj][a].value)
                     #            notEquial.append(wb_b.worksheets[0][jj][a].value)


              #     print(cellobj1)
              #     print(cellobjB1)
    # 创建表
#     new_workbook=xlwt.Workbook()
#     # # 创建sheet
#     worksheet=new_workbook.add_sheet('对比结果')
    
#     for inx,cellobj in enumerate(wb_a.worksheets[0]['D']):
#         if cellobj.value in notEquial:
#     #     # 给对比出来的结果添加颜色高亮
#            cellobj.font = Font(color=colors.BLACK, italic=True ,bold = True)
#            cellobj.fill = PatternFill("solid", fgColor="d71345") 
#     # #  把对比出的结果写入一个附件excel
#            worksheet.write(inx,0,cellobj.value)
#     # #  保存加过对比色文件
#            wb_a.save('/Users/tayli/Desktop/project/python/重标记色全文本对比结果备份.xlsx') 
#     # # 保存晒出掉一样数据的结果，存入excel
#            new_workbook.save('/Users/tayli/Desktop/project/python/只保留不同值excel.xlsx')
      


    return HttpResponse(compare.report(sample_count=5000))