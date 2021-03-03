# -*- coding: UTF-8 -*-
# import numpy as np 
# import pandas as pd    #导入相应的库
# orders = pd.read_excel("/Users/tayli/Desktop/project/python/20200430-三期-修改合同结束时间.xlsx")
# print(orders.head(5))   #读取前五列




import openpyxl
import xlrd

import xlwt
from openpyxl.styles import PatternFill

from openpyxl.styles import colors

from openpyxl.styles import Font, Color

wb_a = openpyxl.load_workbook("/Users/tayli/Desktop/project/python/生效日1.xlsx")

wb_b = openpyxl.load_workbook("/Users/tayli/Desktop/project/python/生效日2.xlsx")

# def getIP(wb):

#     # sheet = wb.get_active_sheet()
#     # get_sheet_by_name('Sheet1')
#     sheet =wb._sheets[0]
# print(wb_a.worksheets[0]['B'][0].value)
# print(wb_a.worksheets[0]['B'][1].value)
notEquial=[]
for i,cellobj in enumerate(wb_a.worksheets[0]['B'],1): 
    # if(wb_a.worksheets[0].value):
    # print(cellobj.value)
    # print(wb_a.worksheets[0]['B'][0].value)
    # 查询b中的某一个伙伴号在a表是不是有,首先伙伴号相等
    # if(cellobj.value==wb_b.worksheets[0]['B'][i].value):
        for j,cellobjB in enumerate(wb_b.worksheets[0]['B'],1):
            # print(cellobj.value,cellobjB)
            if(cellobj.value==cellobjB.value):
                mm=(cellobjB.value,j)
                # 再去判断目标行的值是不是相等 
                if(wb_b.worksheets[0]['D'][j-1].value==wb_a.worksheets[0]['D'][i-1].value):
                        print(wb_a.worksheets[0]['D'][i-1].value)
                else: 
                         print('不相等'+wb_a.worksheets[0]['D'][i-1].value) 
        if(wb_a.worksheets[0]['D'][1].value!=wb_b.worksheets[0]['D'][1].value):
            notEquial.append(wb_b.worksheets[0]['D'][1].value)

# print(wb_b.worksheets[0]['B'][0].value)
